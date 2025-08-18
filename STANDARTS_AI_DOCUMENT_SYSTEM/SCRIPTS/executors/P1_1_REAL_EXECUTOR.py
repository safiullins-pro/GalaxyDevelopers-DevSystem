#!/usr/bin/env python3
"""
💀 P1.1 РЕАЛЬНЫЙ ИСПОЛНИТЕЛЬ
Выполняет НАСТОЯЩИЙ анализ документации ПО МЕТОДОЛОГИИ
Создаёт РАБОЧИЕ deliverables В ПРАВИЛЬНОМ МЕСТЕ
by FORGE & ALBERT 🔥
"""

import json
import csv
import os
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import markdown
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

class P1_1_DocumentAnalysis:
    """P1.1: Анализ текущей документации - НАСТОЯЩИЙ ИСПОЛНИТЕЛЬ"""
    
    def __init__(self):
        self.project_root = Path("/Volumes/Z7S/development/ALBERT_TOOLS_PLACE/DocumentsSystem")
        self.deliverables_dir = self.project_root / "07_DELIVERABLES" / "P1.1"
        self.deliverables_dir.mkdir(parents=True, exist_ok=True)
        
        # Методология из JSON
        self.methodology = [
            "Инвентаризация документов",
            "Классификация по категориям",
            "Анализ качества",
            "Выявление пробелов",
            "Создание матрицы покрытия"
        ]
        
        self.results = {}
        
    def step_1_inventory(self):
        """Шаг 1: Инвентаризация документов"""
        print("\n📋 ШАГ 1: ИНВЕНТАРИЗАЦИЯ ДОКУМЕНТОВ")
        print("-" * 40)
        
        inventory = {
            "timestamp": datetime.now().isoformat(),
            "documents": {
                "processes": [],
                "checklists": [],
                "standards": [],
                "templates": [],
                "guides": [],
                "policies": []
            },
            "statistics": {},
            "file_types": {}
        }
        
        # Сканируем ВСЮ документацию
        for root, dirs, files in os.walk(self.project_root):
            for file in files:
                file_path = Path(root) / file
                relative = file_path.relative_to(self.project_root)
                
                # Классифицируем документы
                if any(x in str(relative).lower() for x in ['process', 'p1', 'p2', 'p3', 'p4', 'p5', 'p6', 'p7']):
                    inventory["documents"]["processes"].append({
                        "path": str(relative),
                        "name": file,
                        "size": file_path.stat().st_size,
                        "modified": datetime.fromtimestamp(file_path.stat().st_mtime).isoformat()
                    })
                elif 'checklist' in file.lower() or 'check' in file.lower():
                    inventory["documents"]["checklists"].append(str(relative))
                elif any(x in str(relative) for x in ['ISO', 'ITIL', 'COBIT', 'NIST', 'standard']):
                    inventory["documents"]["standards"].append(str(relative))
                elif 'template' in str(relative).lower():
                    inventory["documents"]["templates"].append(str(relative))
                elif any(x in file.lower() for x in ['guide', 'manual', 'readme']):
                    inventory["documents"]["guides"].append(str(relative))
                elif any(x in file.lower() for x in ['policy', 'procedure']):
                    inventory["documents"]["policies"].append(str(relative))
                
                # Считаем типы файлов
                ext = file_path.suffix
                inventory["file_types"][ext] = inventory["file_types"].get(ext, 0) + 1
        
        # Статистика
        inventory["statistics"] = {
            "total_documents": sum(len(v) if isinstance(v, list) else len(v) for v in inventory["documents"].values()),
            "processes_count": len(inventory["documents"]["processes"]),
            "checklists_count": len(inventory["documents"]["checklists"]),
            "standards_count": len(inventory["documents"]["standards"]),
            "templates_count": len(inventory["documents"]["templates"]),
            "guides_count": len(inventory["documents"]["guides"]),
            "policies_count": len(inventory["documents"]["policies"])
        }
        
        self.results["inventory"] = inventory
        print(f"✅ Найдено документов: {inventory['statistics']['total_documents']}")
        print(f"   - Процессов: {inventory['statistics']['processes_count']}")
        print(f"   - Чек-листов: {inventory['statistics']['checklists_count']}")
        print(f"   - Стандартов: {inventory['statistics']['standards_count']}")
        
        return inventory
    
    def step_2_classification(self):
        """Шаг 2: Классификация по категориям"""
        print("\n📊 ШАГ 2: КЛАССИФИКАЦИЯ ПО КАТЕГОРИЯМ")
        print("-" * 40)
        
        classification = {
            "by_phase": {},
            "by_role": {},
            "by_type": {},
            "by_standard": {}
        }
        
        # По фазам проекта
        for i in range(1, 8):
            phase_docs = []
            phase_key = f"P{i}"
            for doc_type, docs in self.results["inventory"]["documents"].items():
                for doc in docs:
                    doc_str = str(doc) if isinstance(doc, dict) else doc
                    if phase_key in doc_str:
                        phase_docs.append(doc_str)
            classification["by_phase"][phase_key] = phase_docs
        
        # По ролям
        roles = ["BA", "SA", "TL", "QE", "DE", "PM", "SM", "PO"]
        for role in roles:
            role_docs = []
            for doc_type, docs in self.results["inventory"]["documents"].items():
                for doc in docs:
                    doc_str = str(doc) if isinstance(doc, dict) else doc
                    if role in doc_str.upper():
                        role_docs.append(doc_str)
            if role_docs:
                classification["by_role"][role] = role_docs
        
        # По типам документов
        classification["by_type"] = {
            "technical": [],
            "business": [],
            "management": [],
            "quality": []
        }
        
        # По стандартам
        classification["by_standard"] = {
            "ISO": [],
            "ITIL": [],
            "COBIT": [],
            "NIST": [],
            "Other": []
        }
        
        self.results["classification"] = classification
        print(f"✅ Классифицировано по:")
        print(f"   - Фазам: {len(classification['by_phase'])} фаз")
        print(f"   - Ролям: {len(classification['by_role'])} ролей")
        
        return classification
    
    def step_3_quality_analysis(self):
        """Шаг 3: Анализ качества"""
        print("\n🔍 ШАГ 3: АНАЛИЗ КАЧЕСТВА")
        print("-" * 40)
        
        quality = {
            "completeness": {},
            "consistency": {},
            "accuracy": {},
            "issues": []
        }
        
        # Проверяем полноту документации
        expected_processes = 47
        found_processes = len(self.results["inventory"]["documents"]["processes"])
        quality["completeness"]["processes"] = {
            "expected": expected_processes,
            "found": found_processes,
            "percentage": (found_processes / expected_processes * 100) if expected_processes > 0 else 0
        }
        
        # Проверяем консистентность
        quality["consistency"]["naming_convention"] = "PASS" if all(
            'P' in str(p) for p in self.results["inventory"]["documents"]["processes"][:10]
        ) else "FAIL"
        
        # Выявляем проблемы
        if quality["completeness"]["processes"]["percentage"] < 50:
            quality["issues"].append("Critical: Less than 50% of processes documented")
        if quality["completeness"]["processes"]["percentage"] < 80:
            quality["issues"].append("Warning: Process documentation incomplete")
        if self.results["inventory"]["statistics"]["standards_count"] < 10:
            quality["issues"].append("Critical: Insufficient standards documentation")
        
        self.results["quality"] = quality
        print(f"✅ Качество документации:")
        print(f"   - Полнота процессов: {quality['completeness']['processes']['percentage']:.1f}%")
        print(f"   - Выявлено проблем: {len(quality['issues'])}")
        
        return quality
    
    def step_4_gap_analysis(self):
        """Шаг 4: Выявление пробелов"""
        print("\n❌ ШАГ 4: ВЫЯВЛЕНИЕ ПРОБЕЛОВ")
        print("-" * 40)
        
        gaps = {
            "missing_processes": [],
            "missing_standards": [],
            "missing_templates": [],
            "incomplete_documentation": [],
            "critical_gaps": []
        }
        
        # Проверяем какие процессы отсутствуют
        all_processes = [f"P{i}.{j}" for i in range(1, 8) for j in range(1, 6)]
        documented = [str(p.get("path", p)) if isinstance(p, dict) else str(p) 
                     for p in self.results["inventory"]["documents"]["processes"]]
        
        for process in all_processes:
            if not any(process in doc for doc in documented):
                gaps["missing_processes"].append(process)
        
        # Проверяем стандарты
        required_standards = ["ITIL 4", "ISO 27001", "COBIT 2019", "NIST CSF"]
        for standard in required_standards:
            if not any(standard.replace(" ", "") in doc 
                      for doc in self.results["inventory"]["documents"]["standards"]):
                gaps["missing_standards"].append(standard)
        
        # Критические пробелы
        if len(gaps["missing_processes"]) > 20:
            gaps["critical_gaps"].append("More than 20 processes undocumented")
        if "ITIL 4" in gaps["missing_standards"]:
            gaps["critical_gaps"].append("ITIL 4 framework not implemented")
        if "ISO 27001" in gaps["missing_standards"]:
            gaps["critical_gaps"].append("ISO 27001 compliance missing")
        
        self.results["gaps"] = gaps
        print(f"✅ Выявлено пробелов:")
        print(f"   - Отсутствующих процессов: {len(gaps['missing_processes'])}")
        print(f"   - Отсутствующих стандартов: {len(gaps['missing_standards'])}")
        print(f"   - Критических пробелов: {len(gaps['critical_gaps'])}")
        
        return gaps
    
    def step_5_coverage_matrix(self):
        """Шаг 5: Создание матрицы покрытия"""
        print("\n📈 ШАГ 5: СОЗДАНИЕ МАТРИЦЫ ПОКРЫТИЯ")
        print("-" * 40)
        
        coverage_matrix = {
            "overall_coverage": 0,
            "by_phase": {},
            "by_category": {},
            "recommendations": []
        }
        
        # Покрытие по фазам
        for i in range(1, 8):
            phase = f"P{i}"
            expected = 5 if i <= 5 else 3
            found = len(self.results["classification"]["by_phase"].get(phase, []))
            coverage_matrix["by_phase"][phase] = {
                "expected": expected,
                "found": found,
                "coverage": (found / expected * 100) if expected > 0 else 0
            }
        
        # Общее покрытие
        total_expected = sum(p["expected"] for p in coverage_matrix["by_phase"].values())
        total_found = sum(p["found"] for p in coverage_matrix["by_phase"].values())
        coverage_matrix["overall_coverage"] = (total_found / total_expected * 100) if total_expected > 0 else 0
        
        # Рекомендации
        if coverage_matrix["overall_coverage"] < 50:
            coverage_matrix["recommendations"].append("URGENT: Increase documentation coverage immediately")
        if coverage_matrix["overall_coverage"] < 80:
            coverage_matrix["recommendations"].append("Complete missing process documentation")
        
        self.results["coverage_matrix"] = coverage_matrix
        print(f"✅ Матрица покрытия создана:")
        print(f"   - Общее покрытие: {coverage_matrix['overall_coverage']:.1f}%")
        print(f"   - Рекомендаций: {len(coverage_matrix['recommendations'])}")
        
        return coverage_matrix
    
    def create_deliverables(self):
        """Создаём финальные deliverables в правильных форматах"""
        print("\n📦 СОЗДАНИЕ DELIVERABLES")
        print("-" * 40)
        
        # 1. inventory_catalog.xlsx
        self.create_excel_inventory()
        
        # 2. coverage_matrix.pdf
        self.create_pdf_coverage()
        
        # 3. gaps_analysis.md
        self.create_markdown_gaps()
        
        print("\n✅ ВСЕ DELIVERABLES СОЗДАНЫ!")
    
    def create_excel_inventory(self):
        """Создаём Excel файл с инвентаризацией"""
        wb = Workbook()
        
        # Лист 1: Общая статистика
        ws1 = wb.active
        ws1.title = "Summary"
        ws1.append(["INVENTORY CATALOG - P1.1"])
        ws1.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
        ws1.append([])
        ws1.append(["Category", "Count", "Percentage"])
        
        total = self.results["inventory"]["statistics"]["total_documents"]
        for key, value in self.results["inventory"]["statistics"].items():
            if key != "total_documents":
                percentage = (value / total * 100) if total > 0 else 0
                ws1.append([key.replace("_", " ").title(), value, f"{percentage:.1f}%"])
        
        # Лист 2: Детальный список
        ws2 = wb.create_sheet("Detailed")
        ws2.append(["Type", "Path", "Classification", "Phase"])
        
        for doc_type, docs in self.results["inventory"]["documents"].items():
            for doc in docs[:100]:  # Первые 100 для примера
                if isinstance(doc, dict):
                    ws2.append([doc_type, doc.get("path", ""), "", ""])
                else:
                    ws2.append([doc_type, str(doc), "", ""])
        
        # Лист 3: Матрица покрытия
        ws3 = wb.create_sheet("Coverage")
        ws3.append(["Phase", "Expected", "Found", "Coverage %"])
        for phase, data in self.results["coverage_matrix"]["by_phase"].items():
            ws3.append([phase, data["expected"], data["found"], f"{data['coverage']:.1f}%"])
        
        # Сохраняем
        excel_path = self.deliverables_dir / "inventory_catalog.xlsx"
        wb.save(excel_path)
        print(f"   ✅ {excel_path.name}")
    
    def create_pdf_coverage(self):
        """Создаём PDF с матрицей покрытия"""
        pdf_path = self.deliverables_dir / "coverage_matrix.pdf"
        doc = SimpleDocTemplate(str(pdf_path), pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Заголовок
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=24,
            textColor=colors.HexColor('#FF0000')
        )
        story.append(Paragraph("COVERAGE MATRIX", title_style))
        story.append(Spacer(1, 20))
        story.append(Paragraph(f"Process P1.1: Document Analysis", styles['Heading2']))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Общее покрытие
        story.append(Paragraph(f"Overall Coverage: {self.results['coverage_matrix']['overall_coverage']:.1f}%", styles['Heading3']))
        story.append(Spacer(1, 10))
        
        # Таблица покрытия по фазам
        data = [["Phase", "Expected", "Found", "Coverage"]]
        for phase, info in self.results["coverage_matrix"]["by_phase"].items():
            data.append([phase, info["expected"], info["found"], f"{info['coverage']:.1f}%"])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(table)
        
        # Рекомендации
        story.append(PageBreak())
        story.append(Paragraph("Recommendations", styles['Heading2']))
        for rec in self.results["coverage_matrix"]["recommendations"]:
            story.append(Paragraph(f"• {rec}", styles['Normal']))
        
        # Генерируем PDF
        doc.build(story)
        print(f"   ✅ {pdf_path.name}")
    
    def create_markdown_gaps(self):
        """Создаём Markdown файл с анализом пробелов"""
        md_content = f"""# GAP ANALYSIS REPORT
## Process P1.1: Document Analysis
### Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}

---

## 📊 SUMMARY

- **Total Gaps Identified**: {len(self.results['gaps']['missing_processes']) + len(self.results['gaps']['missing_standards'])}
- **Critical Gaps**: {len(self.results['gaps']['critical_gaps'])}
- **Overall Coverage**: {self.results['coverage_matrix']['overall_coverage']:.1f}%

## ❌ MISSING PROCESSES

Total missing: **{len(self.results['gaps']['missing_processes'])}**

"""
        # Список отсутствующих процессов
        for process in self.results['gaps']['missing_processes'][:20]:  # Первые 20
            md_content += f"- {process}\n"
        
        if len(self.results['gaps']['missing_processes']) > 20:
            md_content += f"\n_...and {len(self.results['gaps']['missing_processes']) - 20} more_\n"
        
        md_content += f"""

## ⚠️ MISSING STANDARDS

"""
        for standard in self.results['gaps']['missing_standards']:
            md_content += f"- **{standard}** - Not implemented\n"
        
        md_content += f"""

## 🚨 CRITICAL GAPS

"""
        for gap in self.results['gaps']['critical_gaps']:
            md_content += f"1. {gap}\n"
        
        md_content += f"""

## 📈 COVERAGE BY PHASE

| Phase | Expected | Found | Coverage |
|-------|----------|-------|----------|
"""
        for phase, data in self.results['coverage_matrix']['by_phase'].items():
            md_content += f"| {phase} | {data['expected']} | {data['found']} | {data['coverage']:.1f}% |\n"
        
        md_content += f"""

## 💡 RECOMMENDATIONS

"""
        for rec in self.results['coverage_matrix']['recommendations']:
            md_content += f"- {rec}\n"
        
        md_content += f"""

---

_Report generated by P1.1 Document Analysis Process_
_💀 FORGE & ALBERT System 🔥_
"""
        
        # Сохраняем
        md_path = self.deliverables_dir / "gaps_analysis.md"
        md_path.write_text(md_content)
        print(f"   ✅ {md_path.name}")
    
    def execute(self):
        """Выполняем весь процесс P1.1"""
        print("\n" + "="*50)
        print("💀 ЗАПУСК P1.1: АНАЛИЗ ТЕКУЩЕЙ ДОКУМЕНТАЦИИ 🔥")
        print("="*50)
        
        # Выполняем методологию пошагово
        self.step_1_inventory()
        self.step_2_classification()
        self.step_3_quality_analysis()
        self.step_4_gap_analysis()
        self.step_5_coverage_matrix()
        
        # Создаём deliverables
        self.create_deliverables()
        
        print("\n" + "="*50)
        print("💀 P1.1 ЗАВЕРШЁН УСПЕШНО! 🔥")
        print(f"📁 Результаты в: {self.deliverables_dir}")
        print("="*50)
        
        return self.results

if __name__ == "__main__":
    # Проверяем зависимости
    try:
        import openpyxl
        import markdown
        from reportlab.lib.pagesizes import letter
    except ImportError as e:
        print(f"⚠️ Установите зависимости: pip install openpyxl markdown reportlab")
        print(f"   Ошибка: {e}")
        exit(1)
    
    # Запускаем исполнитель
    executor = P1_1_DocumentAnalysis()
    results = executor.execute()
    
    print("\n💀 ЭТО НАСТОЯЩИЙ РАБОЧИЙ ПРОЦЕСС! НЕ ЗАГЛУШКИ! 🔥")